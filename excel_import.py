import csv
from openpyxl import load_workbook

workbook_location = 'C:\\Users\\owner\\OneDrive\\Alberta.xlsx'
model_workbook_location = 'C:\\Users\\owner\\OneDrive\\alberta-2023-model.xlsx'
official_candidates = 'official_candidates.csv'
ab_338_ridings_csv = 'ab_338_ridings.csv'
number_of_candidates = 394

def import_official_candidates():
    """
    first build a unique list from the elections alberta list of official candidates
    :return: none
    """
    official = []
    file = open(official_candidates, newline='')
    reader = csv.reader(file, delimiter='\t')
    data = [row for row in reader]
    assert data
    file.close()

    for row in data:
        official.append(row[0])

    wb = load_workbook(workbook_location)
    sheet = wb.worksheets[2]
    count = 0

    for row in range(2, number_of_candidates):
        data = [sheet.cell(row=row, column=i).value for i in range(1, 12)]
        registered = data[5] if data[5] is not None else 'No'
        agent = data[7]
        name = f'{data[1]} {data[2]}'
        if name in official and registered != 'Yes':
            print(f'{name} should be listed')
            count = count + 1
            # sheet.cell(row=row, column=6).value = 'Yes'
        if name in official and registered == 'Yes':
            if agent is None:
                print(f'{name} does not have an official agent listed.')

    print(f'There are {count} candidates that should be listed\n')
    # print('Saving the workbook')
    # wb.save(workbook_location)
    # print('done')

def import_338_riding_info():
    file = open(ab_338_ridings_csv, newline='')
    reader = csv.reader(file, delimiter=',')
    data = [row for row in reader]
    assert data
    file.close()

    wb = load_workbook(model_workbook_location)
    sheet = wb.worksheets[0]
    count = 0

    for row in range(2, 87):
        data = [sheet.cell(row=row, column=i).value for i in range(1, 12)]
        assert False  #todo  add csv fields to excel

    print('excel_import.py: Saving the workbook')
    wb,save(model_workbook_location)

if __name__ == '__main__':
    # import_official_candidates()
    import_338_riding_info()
