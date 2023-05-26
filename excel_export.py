from openpyxl import load_workbook

workbook_location = 'C:\\Users\\owner\\OneDrive\\Alberta.xlsx'


def export_ab_ridings(location=workbook_location):
    wb = load_workbook(workbook_location)
    sheet = wb.worksheets[0]

    with open('ab_ridings.csv', 'wt') as csv_out:
        count = 0
        for row in range(2, 89):
            count = count + 1
            data = [sheet.cell(row=row, column=i).value for i in range(1, 10)]
            csv_out.write(f'{count}\t{data[2]}\t{data[3]}\t{data[4]}\t{data[5]}\t{data[7]}\t{data[8]}\n')
        csv_out.close()


def export_ab_candidates(location=workbook_location, column_headers=True):
    wb = load_workbook(workbook_location)
    sheet = wb.worksheets[2]
    delim = ','
    with open('ab_candidates.csv', 'wt') as csv_out:
        count = 0
        for row in range(1, 400):
            data = [sheet.cell(row=row, column=i).value for i in range(1, 12)]
            short_party = data[0]
            if short_party is not None:
                count = count + 1
                first_name = data[1]
                last_name = data[2]
                riding = data[3]
                incumbent = data[4] if data[4] is not None else 'No'
                registered = data[5] if data[5] is not None else 'No'
                gender = data[6] if data[6] is not None else ''
                year_of_birth = data[7] if data[7] is not None else ''
                url = data[8] if data[8] is not None else ''
                headshot = data[9] if data[9] is not None else ''

                csv_out.write(f'{count}\t{short_party}\t{first_name}\t{last_name}\t{riding}\t{incumbent}\t')
                csv_out.write(f'{registered}\t{gender}\t{year_of_birth}\t{url}\t{headshot}')
                csv_out.write('\n')
                csv_out.flush()
        csv_out.close()
        print(f'excel_export.py: There were {count} official candidates exported.')


def export_candidates_social(location=workbook_location, column_headers=True):
    wb = load_workbook(workbook_location)
    sheet = wb.worksheets[6]
    count = 0
    with open('ab_candidates_social.csv', 'wt') as csv_out:
        for row in range(2, 500):
            data = [sheet.cell(row=row, column=i).value for i in range(1, 10)]
            if data[0] is not None:
                count = count + 1
                csv_out.write(f'{data[0]}\t{data[1]}\t{data[2]}\t{data[3]}\t{data[4]}\t{data[5]}\n')
        csv_out.close()
    print(f'excel_export.py: There were {count} social media links exported.')

def export_ab_parties(location=workbook_location, column_headers=True):
    wb = load_workbook(workbook_location)
    sheet = wb.worksheets[1]
    count = 0
    with open('ab_parties.csv', 'wt') as csv_out:
        for row in range(1, 15):
            count = count + 1
            data = [sheet.cell(row=row, column=i).value for i in range(1, 10)]
            # csv_out.write(f'{count}\t{data[2]}\t{data[3]}\t{data[4]}\t{data[5]}\t{data[7]}\t{data[8]}\n')
            if data[0] is not None:
                csv_out.write(
                    f'{data[0]}\t{data[1]}\t{data[2]}\t{data[3]}\t{data[4]}\t{data[5]}\t{data[6]}\t{data[7]}\t{data[8]}\n')
        csv_out.close()


def export_ab_candidates_nominated_counts(location=workbook_location, column_headers=True):
    wb = load_workbook(workbook_location, data_only=True)
    sheet = wb.worksheets[4]
    count = 0
    with open('ab_candidates_nominated_counts.csv', 'wt') as csv_out:
        for row in range(1, 15):
            count = count + 1
            data = [sheet.cell(row=row, column=i).value for i in range(1, 10)]
            if data[0] is not None:
                csv_out.write(f'{data[0]},{data[1]},{data[2]},{data[3]},{data[4]},{data[5]},{data[6]},{data[7]}\n')
        csv_out.close()
    print('excel_export.py: Exported candidate nominated counts summary')

if __name__ == '__main__':
    # export_ab_ridings()
    # export_ab_parties();
    # export_ab_candidates()
    export_candidates_social()
    # export_ab_candidates_nominated_counts()
