import re
import csv
from openpyxl import load_workbook

workbook_location = 'C:\\Users\\owner\\OneDrive\\Alberta.xlsx'

social_html = 'ndp_candidates_social.html'
csv_file = 'ndp_candidates_social.csv'
social_exp = r'^<li class=\"follow-link-item\">.+?href=\"((https|http)://.+?)\"'
short_name = 'NDP'


def parse_candidates_html():
    file = open(social_html, newline='')
    reader = csv.reader(file, delimiter='\t')
    data = [row for row in reader]
    assert data
    file.close()

    with open(csv_file, 'wt') as csv_out:
        for html in data:
            name = html[0]
            riding = html[1]
            li = html[2]
            link = ''
            platform = ''

            matches = re.match(social_exp, li)
            if matches is not None:
                link = matches[1]
                if 'twitter' in link:
                    platform = "Twitter"
                if 'instagram' in link:
                    platform = 'Instagram'
                if 'facebook' in link:
                    platform = 'Facebook'
            else:
                print(f'{name}: links were not found')
            csv_out.write(f'{short_name}\t{name}\t{riding}\t{platform}\t{link}\n')


def compare_to_excel():
    """
    Compare the list of social media links we extracted from the ndp profiles to the
    list we have in excel.  In excel we may have additional linkls that are not found
    in the local list (ndp_candidates_social.csv). So we can only identify links that
    are not in the excel list.  Also, check if the links are valid.
    :return:
    """

    # first read in the links from excel.
    excel_links = []
    wb = load_workbook(workbook_location)
    sheet = wb.worksheets[6]
    count = 0

    for row in range(2, 500):
        count = count + 1
        data = [sheet.cell(row=row, column=i).value for i in range(1, 10)]
        party = data[0]
        url = str(data[5])

        if party == 'NDP':
            excel_links.append(url.lower())
        else:
            continue

    # print(excel_links)

    file = open(csv_file, newline='')
    reader = csv.reader(file, delimiter='\t')
    data = [row for row in reader]
    assert data
    file.close()

    for csvrow in data:
        url = str(csvrow[4])
        url = url.lower()
        # print(f'checking: {url}')
        if url in excel_links:
            pass
            # print(f'Ok: {url}')
        else:
            print(f'Missing: {url}')


if __name__ == '__main__':
    parse_candidates_html()
    compare_to_excel()
