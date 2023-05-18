import re
import csv
from openpyxl import load_workbook

workbook_location = 'C:\\Users\\owner\\OneDrive\\Alberta.xlsx'
csv_file = 'ucp_candidates_social.csv'

def parse_candidates_html():
    """
    todo: need to copy this code from the more spider. should be here.
    :return:
    """
    pass

def compare_to_excel():
    """
    Compare the list of social media links we extracted from the ndp profiles to the
    list we have in excel.  In excel we may have additional linkls that are not found
    in the local list (ndp_candidates_social.csv). So we can only identify links that
    are not in the excel list.  Also, check if the links are valid.
    :return:
    """
    # first read in the links from excel.
    excel_links = []
    wb = load_workbook(workbook_location)
    sheet = wb.worksheets[5]
    count = 0


    for row in range(2, 400):
        count = count + 1
        data = [sheet.cell(row=row, column=i).value for i in range(1, 10)]
        party = data[0]
        url = str(data[5])

        if party == 'UCP':
            excel_links.append(url.lower())
        else:
            continue



    file = open(csv_file, newline='')
    reader = csv.reader(file, delimiter='\t')
    data = [row for row in reader]
    assert data
    file.close()

    for csvrow in data:
        url = str(csvrow[4])
        url = url.lower()
        if url in excel_links:
            pass
            # print(f'Ok: {url}')
        else:
            print(f'Missing: {url}')



if __name__ == '__main__':
    parse_candidates_html()
    compare_to_excel()